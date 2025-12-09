"""
Script de diagnóstico para revisar por qué se omiten filas en la carga navidad.
Uso:
  cd src
  python scripts/diag_navidad.py "C:\ruta\a\archivo.xlsx" --sheet "" --rows 30

Imprime: fila, fecha parsed, raw_sucursal, normalized_sucursal, store_exists, family_origen, family_exists
"""
import os
import sys
from pathlib import Path
import argparse

if __name__ == '__main__':
    parser = argparse.ArgumentParser()
    parser.add_argument('file', help='Ruta al archivo xlsx')
    parser.add_argument('--sheet', default=None, help='Hoja (nombre o índice)')
    parser.add_argument('--rows', type=int, default=50, help='Cantidad de filas a inspeccionar')
    args = parser.parse_args()

    # Asegurar que la carpeta del proyecto (src) esté en sys.path para poder importar settings
    project_root = Path(__file__).resolve().parent.parent
    if str(project_root) not in sys.path:
        sys.path.insert(0, str(project_root))

    # Configurar Django
    os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'retail_curves.settings')
    try:
        import django
        django.setup()
    except Exception as e:
        print('Error iniciando Django:', e)
        sys.exit(1)

    from core.services.navidad_loader import _read_excel_header, zfill_code, parse_date
    from openpyxl import load_workbook
    from core.models import Store, Family

    p = Path(args.file)
    if not p.exists():
        print('Archivo no encontrado:', p)
        sys.exit(1)

    # Detectar encabezado
    try:
        canon_cols, data_start_row = _read_excel_header(p, args.sheet)
    except Exception as e:
        print('Error detectando encabezado:', e)
        sys.exit(1)

    print('Columnas detectadas:', canon_cols)
    print('Fila inicial de datos (1-based):', data_start_row)

    # Abrir hoja
    wb = load_workbook(p, read_only=True, data_only=True)
    target = (args.sheet if args.sheet not in (None, '') else None)
    if target is None:
        ws = wb.active
    else:
        try:
            target_i = int(target)
            ws = wb.worksheets[target_i]
        except Exception:
            ws = wb[target]

    col_indices = {col_name: idx for idx, col_name in enumerate(canon_cols)}

    inspected = 0
    print('\nInspeccionando primeras', args.rows, 'filas de datos...\n')
    for row_idx, row in enumerate(ws.iter_rows(min_row=data_start_row + 1, values_only=True), start=data_start_row + 1):
        if inspected >= args.rows:
            break
        inspected += 1

        # Construir dict de fila
        row_dict = {col_name: row[col_idx] if col_idx < len(row) else None for col_name, col_idx in col_indices.items()}

        dt = parse_date(row_dict.get('Dia'))
        raw_code = row_dict.get('Sucursal')
        normalized = zfill_code(raw_code, 0)
        subfam = str(row_dict.get('SubFamilia') or '').strip()

        store_exists = Store.objects.filter(code=normalized).exists()
        family_qs = Family.objects.filter(origen=subfam, is_active=True)
        family_exists = family_qs.exists()

        print(f"fila={row_idx} fecha={dt!r} raw_sucursal={raw_code!r} -> normalized={normalized!r} store_exists={store_exists} subfam={subfam!r} family_exists={family_exists}")

    wb.close()
    print('\nTerminado')
