"""
Test rápido: ver cómo openpyxl lee los números con comas.
"""
from openpyxl import load_workbook
from pathlib import Path

p = Path(r'C:\Users\rherrainz\Downloads\Venta para Curvas!(3).xlsx')
wb = load_workbook(p, read_only=True, data_only=True)
ws = wb.active

print('Primeras 10 valores de la columna "Unidades Stock Final":')
for row_idx, row in enumerate(ws.iter_rows(min_row=3, max_row=12, values_only=True), start=3):
    # Asumo que Stock Final es columna índice 5 (0-based, contando desde Dia=0)
    val = row[5] if len(row) > 5 else None
    print(f'  fila {row_idx}: {val!r} (tipo: {type(val).__name__})')

wb.close()
