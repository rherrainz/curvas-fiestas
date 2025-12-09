"""
Diagnóstico: comparar códigos de sucursal en BD vs archivo.
Uso: cd src; python scripts/diag_codes.py "C:\ruta\archivo.xlsx"
"""
import sys
from pathlib import Path
import os

project_root = Path(__file__).resolve().parent.parent
if str(project_root) not in sys.path:
    sys.path.insert(0, str(project_root))
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'retail_curves.settings')
import django
django.setup()

from core.models import Store
from openpyxl import load_workbook
from core.services.navidad_loader import zfill_code

# Códigos en BD
bd_codes = set(Store.objects.values_list('code', flat=True))
print(f'Códigos en BD ({len(bd_codes)}):')
sorted_codes = sorted(bd_codes, key=lambda x: (len(x), x))
for c in sorted_codes[:50]:
    print(f'  {c!r}')
if len(bd_codes) > 50:
    print(f'  ... y {len(bd_codes) - 50} más')

# Códigos en archivo
if len(sys.argv) < 2:
    print('\nUso: python scripts/diag_codes.py "archivo.xlsx"')
    sys.exit(0)

filepath = Path(sys.argv[1])
if not filepath.exists():
    print(f'Archivo no encontrado: {filepath}')
    sys.exit(1)

wb = load_workbook(filepath, read_only=True, data_only=True)
ws = wb.active
file_codes = set()

# Buscar columna "Sucursal" (primera fila)
headers = None
for row_idx, row in enumerate(ws.iter_rows(max_row=5, values_only=True)):
    if headers is None:
        # Buscar "Sucursal"
        row_list = list(row)
        if any('sucursal' in str(v or '').lower() for v in row_list):
            sucursal_idx = next(i for i, v in enumerate(row_list) if 'sucursal' in str(v or '').lower())
            headers = row_idx
            break

if headers is None:
    print('No se encontró columna "Sucursal"')
    sys.exit(1)

# Leer códigos
for row in ws.iter_rows(min_row=headers + 2, values_only=True):
    raw_code = row[sucursal_idx]
    normalized = zfill_code(raw_code, 0)
    if normalized and normalized != '':
        file_codes.add(normalized)

wb.close()

print(f'\nCódigos en archivo ({len(file_codes)}):')
sorted_file = sorted(file_codes, key=lambda x: (len(x), x))
for c in sorted_file[:50]:
    print(f'  {c!r}')
if len(file_codes) > 50:
    print(f'  ... y {len(file_codes) - 50} más')

# Comparar
in_file_not_bd = file_codes - bd_codes
in_bd_not_file = bd_codes - file_codes

print(f'\nEn archivo pero NO en BD ({len(in_file_not_bd)}):')
for c in sorted(in_file_not_bd)[:20]:
    print(f'  {c!r}')
if len(in_file_not_bd) > 20:
    print(f'  ... y {len(in_file_not_bd) - 20} más')

print(f'\nEn BD pero NO en archivo ({len(in_bd_not_file)}):')
for c in sorted(in_bd_not_file)[:20]:
    print(f'  {c!r}')
