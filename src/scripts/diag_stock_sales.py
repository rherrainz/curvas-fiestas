"""
Diagnóstico: inspeccionar parseo de stock vs ventas en primeras filas.
Uso: cd src; python scripts/diag_stock_sales.py "C:\ruta\archivo.xlsx" --rows 20
"""
import sys
from pathlib import Path
import os
import argparse

project_root = Path(__file__).resolve().parent.parent
if str(project_root) not in sys.path:
    sys.path.insert(0, str(project_root))
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'retail_curves.settings')
import django
django.setup()

from core.services.navidad_loader import _read_excel_header, zfill_code, parse_date, parse_number
from openpyxl import load_workbook

parser = argparse.ArgumentParser()
parser.add_argument('file', help='Ruta al archivo xlsx')
parser.add_argument('--rows', type=int, default=20, help='Cantidad de filas a inspeccionar')
args = parser.parse_args()

p = Path(args.file)
if not p.exists():
    print(f'Archivo no encontrado: {p}')
    sys.exit(1)

try:
    canon_cols, data_start_row = _read_excel_header(p, None)
except Exception as e:
    print(f'Error detectando encabezado: {e}')
    sys.exit(1)

print('Columnas detectadas:', canon_cols)
print('Fila inicial de datos:', data_start_row)
print()

# Encontrar índices
col_indices = {col_name: idx for idx, col_name in enumerate(canon_cols)}
print('Índices:')
for col, idx in col_indices.items():
    print(f'  {col}: {idx}')
print()

wb = load_workbook(p, read_only=True, data_only=True)
ws = wb.active

print(f'Primeras {args.rows} filas de datos:\n')
print('fila | fecha | sucursal | subfam | stock_raw | stock_parsed | ventas_raw | ventas_parsed')
print('-' * 100)

inspected = 0
for row_idx, row in enumerate(ws.iter_rows(min_row=data_start_row + 1, values_only=True), start=data_start_row + 1):
    if inspected >= args.rows:
        break
    inspected += 1

    row_dict = {col_name: row[col_idx] if col_idx < len(row) else None for col_name, col_idx in col_indices.items()}
    
    dt = parse_date(row_dict.get('Dia'))
    if not dt:
        continue
    
    code = zfill_code(row_dict.get('Sucursal'), 0)
    subfam = str(row_dict.get('SubFamilia') or '').strip()
    stock_raw = row_dict.get('Unidades Stock Final')
    stock_parsed = parse_number(stock_raw)
    ventas_raw = row_dict.get('Unidades Vendidas')
    ventas_parsed = parse_number(ventas_raw)
    
    print(f'{row_idx:4} | {dt} | {code:8} | {subfam[:15]:15} | {stock_raw!r:9} | {stock_parsed!r:12} | {ventas_raw!r:10} | {ventas_parsed!r:13}')

wb.close()
print('\nTerminado')
