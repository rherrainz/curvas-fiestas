from pathlib import Path
from datetime import date, datetime

import pandas as pd
from openpyxl import load_workbook
from django.db import transaction

from core.models import Store, Family
from sales.models import SalesRecord
from stock.models import StockRecord

REQUIRED_COLS = [
    "Dia", "Region", "Zona", "Sucursal",
    "SubFamilia", "Unidades Stock Final", "Unidades Vendidas",
]


def parse_date(val):
    """
    Normaliza fechas de Excel / strings / datetime a date.
    """
    if val is None:
        return None

    # Si ya viene como Timestamp/datetime/date -> devolvemos date
    if isinstance(val, pd.Timestamp):
        return val.date()
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val

    s = str(val).strip()
    if s == "":
        return None

    # Intento 1: parsear como string día/mes/año
    try:
        return pd.to_datetime(s, dayfirst=True, errors="raise").date()
    except Exception:
        pass

    # Intento 2: manejar números de fecha de Excel
    try:
        return pd.to_datetime(float(s), unit="D", origin="1899-12-30").date()
    except Exception:
        return None


def parse_number(val):
    """
    Normaliza números con coma/punto/espacios a float o None.
    """
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return None
    if isinstance(val, (int, float)):
        return float(val)

    s = str(val).strip().replace(" ", "")
    if s == "":
        return None

    if "," in s and "." in s:
        last_comma, last_dot = s.rfind(","), s.rfind(".")
        if last_comma > last_dot:
            s = s.replace(".", "").replace(",", ".")
        else:
            s = s.replace(",", "")
    else:
        if "," in s:
            s = s.replace(".", "").replace(",", ".")

    try:
        return float(s)
    except Exception:
        return None


def zfill_code(raw, pad: int) -> str:
    """
    Normaliza el código de sucursal. Hoy ignora 'pad' a propósito
    (no agrega ceros, solo limpia). Si querés realmente pad-left,
    acá es donde tocarías.
    """
    s = str(raw).strip()
    try:
        num = float(s)
        if num.is_integer():
            s = str(int(num))
    except Exception:
        pass

    if s.isdigit():
        s = s.lstrip("0") or "0"
    return s


def _detect_header_row(df_no_header: pd.DataFrame, max_scan: int = 30):
    def norm(s):
        s = "" if s is None else str(s)
        return " ".join(s.replace("\n", " ").strip().lower().split())

    required = [
        "Dia", "Region", "Zona", "Sucursal",
        "SubFamilia", "Unidades Stock Final", "Unidades Vendidas",
    ]
    aliases = {
        "día": "Dia", "dia": "Dia", "fecha": "Dia",
        "region": "Region", "región": "Region",
        "zona": "Zona",
        "sucursal": "Sucursal", "tienda": "Sucursal", "store": "Sucursal",
        "subfamilia": "SubFamilia", "sub familia": "SubFamilia", "origen": "SubFamilia",
        "unidades stock final": "Unidades Stock Final", "stock final unidades": "Unidades Stock Final",
        "stock final": "Unidades Stock Final", "stock (unidades)": "Unidades Stock Final",
        "unidades vendidas": "Unidades Vendidas", "ventas unidades": "Unidades Vendidas", "ventas": "Unidades Vendidas",
    }

    required_norm = [norm(x) for x in required]
    aliases_norm = {k: v for k, v in aliases.items()}

    best = (-1, -1, None)
    nrows = min(len(df_no_header), max_scan)

    for i in range(nrows):
        row = list(df_no_header.iloc[i].values)
        cols_map = []
        matches = 0

        for val in row:
            key = norm(val)
            canon = None
            if key in aliases_norm:
                canon = aliases_norm[key]
            elif key in required_norm:
                canon = required[required_norm.index(key)]

            cols_map.append(canon)
            if canon in required:
                matches += 1

        if matches > best[0]:
            best = (matches, i, cols_map)
        if matches == len(required):
            break

    if best[0] >= 5:
        return best[1], best[2]
    return None, None


def _read_excel_header(path: Path, sheet: str | None) -> tuple[list[str], int]:
    """
    Lee solo la cabecera del Excel (primeras ~100 filas) y retorna:
    (nombres_canonicos, fila_datos_inicio)
    """
    p = Path(path)
    target = (sheet if sheet not in (None, "") else 0)

    x = pd.read_excel(p, sheet_name=target, header=None, nrows=100)
    if isinstance(x, dict):
        raw = x[target] if isinstance(target, str) and target in x else next(iter(x.values()))
    else:
        raw = x

    hdr_idx, cols_map = _detect_header_row(raw, max_scan=30)
    if hdr_idx is None:
        raise ValueError("No pude detectar la fila de encabezados. Verificá el archivo/hoja.")

    header_row_values = list(raw.iloc[hdr_idx].values)
    canon_cols: list[str] = []

    for j, v in enumerate(header_row_values):
        canon = cols_map[j]
        if canon is None:
            canon = str(v).strip() if v is not None else f"col_{j}"
        canon_cols.append(canon)

    miss = [c for c in REQUIRED_COLS if c not in canon_cols]
    if miss:
        raise ValueError(f"Faltan columnas requeridas: {', '.join(miss)}")

    # hdr_idx + 1 es donde empiezan los datos (fila siguiente a encabezados)
    return canon_cols, hdr_idx + 1


@transaction.atomic
def process_navidad_file(
    path: Path,
    *,
    sheet: str | None,
    pad: int = 0,
    strict_area: bool = False,
    chunk_size: int = 10_000,
):
    """
    Loader masivo para archivo de Navidad:
    - Detecta encabezados automáticamente.
    - Lee con openpyxl para no cargar todo en memoria.
    - Procesa en chunks y hace bulk_create / bulk_update.
    """
    p = Path(path)
    if not p.exists():
        raise FileNotFoundError(p)

    canon_cols, data_start_row = _read_excel_header(p, sheet)

    summary = {
        "stock_created": 0,
        "stock_updated": 0,
        "stock_skipped": 0,
        "sales_created": 0,
        "sales_updated": 0,
        "sales_skipped": 0,
        "rows": 0,
        "detected_columns": canon_cols,
        "rows_raw": 0,
    }

    store_cache: dict[str, Store | None] = {}
    family_cache: dict[tuple[str, str], Family | None] = {}

    target_sheet = (sheet if sheet not in (None, "") else None)
    wb = load_workbook(p, read_only=True, data_only=True)

    if target_sheet is None:
        ws = wb.active
    elif isinstance(target_sheet, int):
        ws = wb.worksheets[target_sheet]
    else:
        ws = wb[target_sheet]

    col_indices = {col_name: idx for idx, col_name in enumerate(canon_cols)}

    try:
        total_rows = max(0, ws.max_row - data_start_row)
    except Exception:
        total_rows = None

    report_every = max(1, min(1000, chunk_size))

    if total_rows:
        print(f"[navidad_loader] Procesando '{p.name}' hoja='{sheet}' filas_aprox={total_rows}", flush=True)
    else:
        print(f"[navidad_loader] Procesando '{p.name}' hoja='{sheet}' (filas totales desconocidas)", flush=True)

    row_count = 0
    chunk: list[dict] = []

    def flush_chunk(chunk_rows: list[dict]):
        nonlocal summary
        if not chunk_rows:
            return

        codes = set()
        subfams = set()
        dates = set()

        for r in chunk_rows:
            codes.add(r["code"])
            subfams.add(r["subfam"])
            dates.add(r["date"])

        codes_list = list(codes)
        subfams_list = list(subfams)
        dates_list = list(dates)

        # Pre-cargar Stores en cache
        stores_qs = Store.objects.filter(code__in=codes_list)
        for s in stores_qs:
            store_cache[s.code] = s

        # Pre-cargar Families activas en cache (por origen)
        families_qs = Family.objects.filter(origen__in=subfams_list, is_active=True)
        for f in families_qs:
            family_cache[(f.origen, "is_active")] = f

        store_ids = [
            store_cache[c].id
            for c in codes_list
            if store_cache.get(c) is not None
        ]
        family_ids = [
            family_cache.get((sf, "is_active")).id
            for sf in subfams_list
            if family_cache.get((sf, "is_active")) is not None
        ]

        existing_stock = {}
        existing_sales = {}

        if store_ids and family_ids and dates_list:
            stock_qs = StockRecord.objects.filter(
                store_id__in=store_ids,
                family_id__in=family_ids,
                date__in=dates_list,
            )
            for obj in stock_qs:
                existing_stock[(obj.store_id, obj.family_id, obj.date)] = obj

            sales_qs = SalesRecord.objects.filter(
                store_id__in=store_ids,
                family_id__in=family_ids,
                date__in=dates_list,
            )
            for obj in sales_qs:
                existing_sales[(obj.store_id, obj.family_id, obj.date)] = obj

        to_create_stock: list[StockRecord] = []
        to_update_stock: list[StockRecord] = []
        to_create_sales: list[SalesRecord] = []
        to_update_sales: list[SalesRecord] = []

        for r in chunk_rows:
            store = store_cache.get(r["code"])
            if store is None:
                summary["stock_skipped"] += 1
                summary["sales_skipped"] += 1
                continue

            if strict_area:
                region_excel = r.get("region", "")
                zona_excel = r.get("zona", "")
                try:
                    if (
                        store.region.name.strip() != region_excel
                        or store.zone.name.strip() != zona_excel
                    ):
                        summary["stock_skipped"] += 1
                        summary["sales_skipped"] += 1
                        continue
                except Exception:
                    summary["stock_skipped"] += 1
                    summary["sales_skipped"] += 1
                    continue

            family = family_cache.get((r["subfam"], "is_active"))
            if family is None:
                summary["stock_skipped"] += 1
                summary["sales_skipped"] += 1
                continue

            # STOCK
            if r["stock_units"] is not None:
                key = (store.id, family.id, r["date"])
                existing = existing_stock.get(key)
                if existing:
                    existing.stock_units = r["stock_units"]
                    existing.stock_value = None
                    to_update_stock.append(existing)
                else:
                    to_create_stock.append(
                        StockRecord(
                            store=store,
                            family=family,
                            date=r["date"],
                            stock_units=r["stock_units"],
                            stock_value=None,
                        )
                    )
            else:
                summary["stock_skipped"] += 1

            # SALES
            if (
                r["units_sold"] is not None
                and r["units_sold"] > 0
                and not store.is_distribution_center
            ):
                key = (store.id, family.id, r["date"])
                existing = existing_sales.get(key)
                if existing:
                    existing.units_sold = r["units_sold"]
                    existing.revenue = None
                    to_update_sales.append(existing)
                else:
                    to_create_sales.append(
                        SalesRecord(
                            store=store,
                            family=family,
                            date=r["date"],
                            units_sold=r["units_sold"],
                            revenue=None,
                        )
                    )
            else:
                summary["sales_skipped"] += 1

        # Transacción por chunk (anidada a la global)
        with transaction.atomic():
            if to_create_stock:
                StockRecord.objects.bulk_create(to_create_stock)
                summary["stock_created"] += len(to_create_stock)
            if to_update_stock:
                StockRecord.objects.bulk_update(
                    to_update_stock,
                    ["stock_units", "stock_value"],
                )
                summary["stock_updated"] += len(to_update_stock)
            if to_create_sales:
                SalesRecord.objects.bulk_create(to_create_sales)
                summary["sales_created"] += len(to_create_sales)
            if to_update_sales:
                SalesRecord.objects.bulk_update(
                    to_update_sales,
                    ["units_sold", "revenue"],
                )
                summary["sales_updated"] += len(to_update_sales)

        # Debug del chunk
        stock_none_count = sum(
            1 for r in chunk_rows if r["stock_units"] is None
        )
        print(
            f"[chunk_flush] stock: crear={len(to_create_stock)} actualizar={len(to_update_stock)} "
            f"sin_valor={stock_none_count} | sales: crear={len(to_create_sales)} "
            f"actualizar={len(to_update_sales)}",
            flush=True,
        )

    # Bucle principal de filas Excel
    for row_idx, row in enumerate(
        ws.iter_rows(min_row=data_start_row + 1, values_only=True),
        start=data_start_row + 1,
    ):
        row_count += 1

        if row_count % report_every == 0:
            if total_rows:
                pct = (row_count / total_rows) * 100
                print(
                    f"[navidad_loader] Procesadas {row_count}/{total_rows} filas ({pct:.1f}%)",
                    flush=True,
                )
            else:
                print(
                    f"[navidad_loader] Procesadas {row_count} filas...",
                    flush=True,
                )

        summary["rows"] += 1
        summary["rows_raw"] += 1

        row_dict = {
            col_name: row[col_idx] if col_idx < len(row) else None
            for col_name, col_idx in col_indices.items()
        }

        dt = parse_date(row_dict["Dia"])
        if not dt:
            continue

        code = zfill_code(row_dict["Sucursal"], pad)
        subfam_excel = str(row_dict["SubFamilia"] or "").strip()
        if not subfam_excel:
            continue

        stock_units = parse_number(row_dict.get("Unidades Stock Final"))
        units_sold = parse_number(row_dict.get("Unidades Vendidas"))

        chunk.append(
            {
                "date": dt,
                "code": code,
                "subfam": subfam_excel,
                "stock_units": stock_units,
                "units_sold": units_sold,
                "region": str(row_dict.get("Region") or "").strip(),
                "zona": str(row_dict.get("Zona") or "").strip(),
            }
        )

        if len(chunk) >= chunk_size:
            flush_chunk(chunk)
            chunk = []

    if chunk:
        flush_chunk(chunk)

    wb.close()

    if total_rows:
        print(
            f"[navidad_loader] Procesado final: {row_count}/{total_rows} filas",
            flush=True,
        )
    else:
        print(
            f"[navidad_loader] Procesado final: {row_count} filas",
            flush=True,
        )

    return summary

